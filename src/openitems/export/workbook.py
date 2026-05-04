"""Open Items List → .xlsx writer.

Direct port of the VBA `BuildReport` flow (`modOpenItemsList.bas:369-793`).
Output columns A–I:
  A=spacer, B=#, C=Task, D=Tags, E=Priority, F=Assigned To,
  G=Start, H=Due, I=Description / Checklist
"""

from __future__ import annotations

import math
from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from openitems.db.models import Engagement, Policy, Task
from openitems.domain.policies import days_to_renewal
from openitems.domain.tasks import (
    high_priority_count,
    is_completed,
    is_late,
    overdue_count,
    total_checks,
)
from openitems.domain.text import clean_text
from openitems.export.theme import (
    CLR_CHARCOAL,
    CLR_CHK_BG,
    CLR_CREAM,
    CLR_GRAY_LT,
    CLR_GRAY_MED,
    CLR_NAVY,
    CLR_RED,
    CLR_SUBTOTAL,
    CLR_WHITE,
    FONT_NAME,
)

# A=1 spacer, B=2 #, C=3 Task, D=4 Tags, E=5 Pri, F=6 Asg, G=7 Start, H=8 Due, I=9 Desc
OUT_COLS = 9
COL_WIDTHS: dict[str, float] = {
    "A": 2.5,
    "B": 5,
    "C": 42,
    "D": 14,
    "E": 11,
    "F": 22,
    "G": 14,
    "H": 14,
    "I": 50,
}

HEADERS = ("", "#", "Task", "Tags", "Priority", "Assigned To", "Start", "Due", "Description / Checklist")

# Approximate per-character column capacity for wrapped-text auto-sizing.
# Excel's column width is in "characters of the default font," but with
# proportional fonts and padding the effective wrap width runs ~10% lower.
_WRAP_FUDGE = 0.92
_LINE_HEIGHT_PT = 15.0  # ~ font size 10 + leading
_MIN_TASK_ROW_HEIGHT = 22.0


def _fill(rgb: str) -> PatternFill:
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")


def _apply_band(ws: Worksheet, row: int, rgb: str) -> None:
    f = _fill(rgb)
    for c in range(1, OUT_COLS + 1):
        ws.cell(row=row, column=c).fill = f


def _format_date(d: date | None) -> str:
    if d is None:
        return ""
    return d.strftime("%m-%d-%Y")


def _is_overdue(t: Task, today: date) -> bool:
    return is_late(t, today)


def _bottom_border(color: str = CLR_GRAY_LT, weight: str = "thin") -> Border:
    return Border(
        bottom=Side(border_style=weight, color=color),
    )


def _wrapped_lines_for_width(text: str | None, width: float) -> int:
    """Estimate visual line count when wrapped at column ``width`` chars."""
    if not text:
        return 1
    capacity = max(1.0, width * _WRAP_FUDGE)
    total = 0
    for raw in str(text).splitlines() or [""]:
        if not raw:
            total += 1
        else:
            total += max(1, math.ceil(len(raw) / capacity))
    return max(1, total)


def _wrapped_lines(text: str | None, col_letter: str) -> int:
    """Estimate the number of visual lines `text` occupies in column `col_letter`."""
    return _wrapped_lines_for_width(text, COL_WIDTHS[col_letter])


def _row_height_for(*texts_and_cols: tuple[str | None, str]) -> float:
    lines = max(_wrapped_lines(t, c) for t, c in texts_and_cols)
    return max(_MIN_TASK_ROW_HEIGHT, lines * _LINE_HEIGHT_PT + 6)


def _write_title_block(
    ws: Worksheet,
    *,
    client_name: str,
    total_open: int,
    bucket_count: int,
) -> None:
    # Row 1 — client name
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 28
    client = ws.cell(row=1, column=1, value=f"  {client_name}")
    client.font = Font(name=FONT_NAME, size=12, bold=True, color=CLR_WHITE)
    client.alignment = Alignment(vertical="center")
    _apply_band(ws, 1, CLR_NAVY)

    # Row 2 — title
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 52
    title = ws.cell(row=2, column=1, value="  Open Items List")
    title.font = Font(name=FONT_NAME, size=22, bold=True, color=CLR_WHITE)
    title.alignment = Alignment(vertical="center")
    _apply_band(ws, 2, CLR_NAVY)

    # Row 3 — subtitle
    ws.merge_cells("A3:I3")
    ws.row_dimensions[3].height = 22
    subtitle_text = (
        "  Generated: "
        + datetime.now().strftime("%B %-d, %Y %-I:%M %p")
        + f"  |  {total_open} open items  |  {bucket_count} buckets"
    )
    sub = ws.cell(row=3, column=1, value=subtitle_text)
    sub.font = Font(name=FONT_NAME, size=9, color=CLR_GRAY_MED)
    sub.alignment = Alignment(vertical="center")
    _apply_band(ws, 3, CLR_SUBTOTAL)

    # Row 4 — spacer
    ws.row_dimensions[4].height = 6

    # Row 5 — column headers
    for c, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=5, column=c, value=header)
        cell.font = Font(name=FONT_NAME, size=9, bold=True, color=CLR_WHITE)
        cell.alignment = Alignment(
            vertical="center",
            horizontal="center" if c <= 6 else "general",
        )
        cell.fill = _fill(CLR_CHARCOAL)
    ws.row_dimensions[5].height = 26


def _write_bucket_header(ws: Worksheet, row: int, bucket_name: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=OUT_COLS)
    ws.row_dimensions[row].height = 28
    cell = ws.cell(row=row, column=1, value=f"  {bucket_name.upper()}")
    cell.font = Font(name=FONT_NAME, size=12, bold=True, color=CLR_WHITE)
    cell.alignment = Alignment(vertical="center")
    _apply_band(ws, row, CLR_NAVY)


def _write_task_row(
    ws: Worksheet, row: int, task: Task, fill_rgb: str, today: date
) -> None:
    f = _fill(fill_rgb)
    overdue = _is_overdue(task, today)
    high = task.priority in {"Urgent", "Important"}
    description = clean_text(task.description)

    # A: spacer
    ws.cell(row=row, column=1).fill = f

    # B: #
    cell = ws.cell(row=row, column=2)
    cell.font = Font(name=FONT_NAME, size=10, color=CLR_GRAY_MED)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = f

    # C: Task name
    cell = ws.cell(row=row, column=3, value=task.name)
    cell.font = Font(
        name=FONT_NAME, size=10, color=CLR_CHARCOAL, bold=overdue
    )
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.fill = f

    # D: Tags
    cell = ws.cell(row=row, column=4, value=task.labels)
    cell.font = Font(name=FONT_NAME, size=10, color=CLR_CHARCOAL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = f

    # E: Priority
    cell = ws.cell(row=row, column=5, value=task.priority)
    cell.font = Font(
        name=FONT_NAME,
        size=10,
        color=CLR_RED if high else CLR_CHARCOAL,
        bold=high,
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = f

    # F: Assigned
    cell = ws.cell(row=row, column=6, value=task.assigned_to)
    cell.font = Font(name=FONT_NAME, size=10, color=CLR_CHARCOAL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = f

    # G: Start
    cell = ws.cell(row=row, column=7, value=_format_date(task.start_date))
    cell.font = Font(name=FONT_NAME, size=10, color=CLR_CHARCOAL)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = f

    # H: Due
    cell = ws.cell(row=row, column=8, value=_format_date(task.due_date))
    cell.font = Font(
        name=FONT_NAME,
        size=10,
        color=CLR_RED if overdue else CLR_CHARCOAL,
        bold=overdue,
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = f

    # I: Description
    cell = ws.cell(row=row, column=9, value=description)
    cell.font = Font(name=FONT_NAME, size=10, color=CLR_CHARCOAL)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.fill = f

    # Bottom hairline border
    border = _bottom_border()
    for c in range(1, OUT_COLS + 1):
        ws.cell(row=row, column=c).border = border

    # Auto-size row to the tallest wrapped column (Task or Description).
    ws.row_dimensions[row].height = _row_height_for(
        (task.name, "C"), (description, "I")
    )


def _write_checklist_block(ws: Worksheet, start_row: int, task: Task) -> int:
    """Render checklist progress + items beneath the task row.

    Returns the next free row index.
    """
    items = [c for c in task.checklist_items if c.deleted_at is None]
    if not items:
        return start_row

    completed = sum(1 for c in items if c.completed)
    total = len(items)

    # Progress label row
    row = start_row
    ws.row_dimensions[row].height = 18
    _apply_band(ws, row, CLR_CHK_BG)
    cell = ws.cell(row=row, column=3, value=f"    Checklist: {completed} of {total} complete")
    cell.font = Font(name=FONT_NAME, size=9, italic=True, color=CLR_GRAY_MED)
    cell.alignment = Alignment(indent=2, vertical="center")

    # Item rows. We render in the stored order rather than recomputing
    # completed-first — the user's ordering is meaningful.
    border = Border(
        bottom=Side(border_style="hair", color="D2D7DC"),
    )
    for item in items:
        row += 1
        _apply_band(ws, row, CLR_CHK_BG)
        for c in range(1, OUT_COLS + 1):
            ws.cell(row=row, column=c).border = border
        prefix = "[x]  " if item.completed else "[  ]  "
        cell = ws.cell(row=row, column=3, value=prefix + item.text)
        cell.font = Font(name=FONT_NAME, size=9, italic=True, color=CLR_GRAY_MED)
        cell.alignment = Alignment(indent=3, wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = max(
            18.0, _wrapped_lines(prefix + item.text, "C") * _LINE_HEIGHT_PT + 4
        )

    return row + 1


def _write_summary(
    ws: Worksheet,
    row: int,
    *,
    total_open: int,
    bucket_count: int,
    total_high: int,
    total_overdue: int,
) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=OUT_COLS)
    ws.row_dimensions[row].height = 24
    cell = ws.cell(
        row=row,
        column=1,
        value=(
            f"  Summary: {total_open} open items across {bucket_count} buckets"
            f"  |  {total_high} high/urgent priority"
            f"  |  {total_overdue} overdue"
        ),
    )
    cell.font = Font(name=FONT_NAME, size=9, bold=True, color=CLR_NAVY)
    cell.alignment = Alignment(vertical="center")
    _apply_band(ws, row, CLR_SUBTOTAL)


def _apply_page_setup(ws: Worksheet, last_row: int) -> None:
    ws.print_area = f"A1:{get_column_letter(OUT_COLS)}{last_row}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # unlimited vertical pages
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.header = 0.25
    ws.page_margins.footer = 0.25
    ws.print_options.horizontalCentered = True
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"


def export_engagement(
    engagement: Engagement,
    tasks: Iterable[Task],
    output_path: Path,
    *,
    today: date | None = None,
    policies: Iterable[Policy] | None = None,
) -> Path:
    """Write a Planner-style Open Items List for `engagement`.

    A second ``Policies`` tab is appended when ``policies`` contains at
    least one live row — keeping single-engagement clients without a
    policy list on a clean one-tab workbook.
    """
    today = today or date.today()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    open_tasks = [t for t in tasks if t.deleted_at is None and not is_completed(t)]

    # Group by bucket name. Order by workflow sort_order (Backlog → In Progress
    # → In Review …), falling back to alphabetical for any pre-refactor
    # buckets that share a sort_order. This is the one place the new export
    # diverges from the original VBA `:223` (which sorted alphabetically).
    by_bucket: dict[str, list[Task]] = {}
    bucket_meta: dict[str, tuple[int, str]] = {}
    for t in open_tasks:
        if t.bucket:
            bucket_name = t.bucket.name
            bucket_meta[bucket_name] = (t.bucket.sort_order, bucket_name)
        else:
            bucket_name = "(no bucket)"
            bucket_meta.setdefault(bucket_name, (10_000, bucket_name))
        by_bucket.setdefault(bucket_name, []).append(t)
    bucket_names = sorted(by_bucket.keys(), key=lambda n: bucket_meta[n])

    total_open = len(open_tasks)
    total_overdue = overdue_count(open_tasks, today)
    total_high = high_priority_count(open_tasks)

    wb = Workbook()
    ws = wb.active
    sheet_name = f"Open Items - {today.strftime('%Y-%m-%d')}"
    ws.title = sheet_name[:31]  # Excel limit
    ws.sheet_properties.tabColor = CLR_NAVY

    # Bucket headers stay visible while the rows beneath them collapse.
    ws.sheet_properties.outlinePr.summaryBelow = False

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    _write_title_block(
        ws,
        client_name=engagement.name,
        total_open=total_open,
        bucket_count=len(bucket_names),
    )

    row = 6
    for bucket_name in bucket_names:
        _write_bucket_header(ws, row, bucket_name)
        bucket_first_row = row + 1
        for idx, task in enumerate(by_bucket[bucket_name], start=1):
            row += 1
            fill_rgb = CLR_CREAM if idx % 2 == 0 else CLR_WHITE
            _write_task_row(ws, row, task, fill_rgb, today)
            # Number cell after we know the placement so it matches running count
            ws.cell(row=row, column=2, value=_running_index(by_bucket, bucket_names, bucket_name, idx))
            # Checklist sub-rows (if any)
            if total_checks(task) > 0:
                row = _write_checklist_block(ws, row + 1, task) - 1
        # Group every row beneath the bucket header so the bucket is collapsible.
        for r in range(bucket_first_row, row + 1):
            ws.row_dimensions[r].outline_level = 1
        row += 1
        ws.row_dimensions[row].height = 8  # spacer
        row += 1

    _write_summary(
        ws,
        row,
        total_open=total_open,
        bucket_count=len(bucket_names),
        total_high=total_high,
        total_overdue=total_overdue,
    )

    _apply_page_setup(ws, row)

    live_policies = (
        [p for p in policies if p.deleted_at is None and p.archived_at is None]
        if policies
        else []
    )
    if live_policies:
        live_policies.sort(
            key=lambda p: (
                p.expiration_date is None,
                p.expiration_date or date.max,
                (p.carrier or "").casefold(),
                p.name.casefold(),
            )
        )
        _write_policies_sheet(
            wb, engagement=engagement, policies=live_policies, today=today
        )

    wb.save(output_path)
    return output_path


_POL_HEADERS = (
    "",
    "#",
    "Policy",
    "Carrier",
    "Coverage",
    "Policy #",
    "Location",
    "Effective",
    "Expiration",
    "Days to Renewal",
    "Description",
)
_POL_COL_WIDTHS: dict[str, float] = {
    "A": 2.5,
    "B": 5,
    "C": 28,
    "D": 18,
    "E": 14,
    "F": 16,
    "G": 22,
    "H": 14,
    "I": 14,
    "J": 16,
    "K": 36,
}
_POL_OUT_COLS = len(_POL_HEADERS)


def _write_policies_sheet(
    wb: Workbook,
    *,
    engagement: Engagement,
    policies: list[Policy],
    today: date,
) -> None:
    sheet_name = f"Policies - {today.strftime('%Y-%m-%d')}"
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.sheet_properties.tabColor = CLR_NAVY

    for col, width in _POL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Title block — mirrors the Open Items sheet for visual continuity.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_POL_OUT_COLS)
    ws.row_dimensions[1].height = 28
    client = ws.cell(row=1, column=1, value=f"  {engagement.name}")
    client.font = Font(name=FONT_NAME, size=12, bold=True, color=CLR_WHITE)
    client.alignment = Alignment(vertical="center")
    for c in range(1, _POL_OUT_COLS + 1):
        ws.cell(row=1, column=c).fill = _fill(CLR_NAVY)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_POL_OUT_COLS)
    ws.row_dimensions[2].height = 52
    title = ws.cell(row=2, column=1, value="  Policies")
    title.font = Font(name=FONT_NAME, size=22, bold=True, color=CLR_WHITE)
    title.alignment = Alignment(vertical="center")
    for c in range(1, _POL_OUT_COLS + 1):
        ws.cell(row=2, column=c).fill = _fill(CLR_NAVY)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=_POL_OUT_COLS)
    ws.row_dimensions[3].height = 22
    lapsed = sum(1 for p in policies if (d := days_to_renewal(p, today)) is not None and d < 0)
    soon = sum(
        1
        for p in policies
        if (d := days_to_renewal(p, today)) is not None and 0 <= d <= 30
    )
    subtitle = ws.cell(
        row=3,
        column=1,
        value=(
            "  Generated: "
            + datetime.now().strftime("%B %-d, %Y %-I:%M %p")
            + f"  |  {len(policies)} policies"
            + (f"  |  {lapsed} lapsed" if lapsed else "")
            + (f"  |  {soon} renewing ≤30d" if soon else "")
        ),
    )
    subtitle.font = Font(name=FONT_NAME, size=9, color=CLR_GRAY_MED)
    subtitle.alignment = Alignment(vertical="center")
    for c in range(1, _POL_OUT_COLS + 1):
        ws.cell(row=3, column=c).fill = _fill(CLR_SUBTOTAL)

    ws.row_dimensions[4].height = 6  # spacer

    for c, header in enumerate(_POL_HEADERS, start=1):
        cell = ws.cell(row=5, column=c, value=header)
        cell.font = Font(name=FONT_NAME, size=9, bold=True, color=CLR_WHITE)
        cell.alignment = Alignment(
            vertical="center", horizontal="center" if c <= 6 else "general"
        )
        cell.fill = _fill(CLR_CHARCOAL)
    ws.row_dimensions[5].height = 26

    border = _bottom_border()
    row = 5
    for idx, p in enumerate(policies, start=1):
        row += 1
        d = days_to_renewal(p, today)
        lapsed_row = d is not None and d < 0
        soon_row = d is not None and 0 <= d <= 30
        fill_rgb = CLR_CREAM if idx % 2 == 0 else CLR_WHITE
        f = _fill(fill_rgb)

        for c in range(1, _POL_OUT_COLS + 1):
            ws.cell(row=row, column=c).fill = f

        ws.cell(row=row, column=2, value=idx).font = Font(
            name=FONT_NAME, size=10, color=CLR_GRAY_MED
        )
        ws.cell(row=row, column=2).alignment = Alignment(
            horizontal="center", vertical="center"
        )

        cells = (
            (3, p.name, True),
            (4, p.carrier, False),
            (5, p.coverage, False),
            (6, p.policy_number, False),
            (7, p.location, False),
            (8, _format_date(p.effective_date), False),
        )
        for col, value, bold_when_lapsed in cells:
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(
                name=FONT_NAME,
                size=10,
                color=CLR_CHARCOAL,
                bold=bold_when_lapsed and lapsed_row,
            )
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        # Two visual tiers per the spec: bold+red for lapsed (loudest), red
        # only for ≤30d (a quieter warning), charcoal otherwise.
        exp_cell = ws.cell(row=row, column=9, value=_format_date(p.expiration_date))
        if lapsed_row:
            exp_cell.font = Font(name=FONT_NAME, size=10, color=CLR_RED, bold=True)
        elif soon_row:
            exp_cell.font = Font(name=FONT_NAME, size=10, color=CLR_RED, bold=False)
        else:
            exp_cell.font = Font(name=FONT_NAME, size=10, color=CLR_CHARCOAL)
        exp_cell.alignment = Alignment(horizontal="center", vertical="center")

        days_value = (
            "—" if d is None else (f"lapsed {-d}d" if d < 0 else f"{d}d")
        )
        days_cell = ws.cell(row=row, column=10, value=days_value)
        if lapsed_row:
            days_cell.font = Font(name=FONT_NAME, size=10, color=CLR_RED, bold=True)
        elif soon_row:
            days_cell.font = Font(name=FONT_NAME, size=10, color=CLR_RED, bold=False)
        else:
            days_cell.font = Font(name=FONT_NAME, size=10, color=CLR_GRAY_MED)
        days_cell.alignment = Alignment(horizontal="center", vertical="center")

        desc_cell = ws.cell(row=row, column=11, value=p.description or "")
        desc_cell.font = Font(name=FONT_NAME, size=10, color=CLR_CHARCOAL)
        desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

        for c in range(1, _POL_OUT_COLS + 1):
            ws.cell(row=row, column=c).border = border

        # Auto-size to whichever wrapped column is tallest.
        ws.row_dimensions[row].height = max(
            _MIN_TASK_ROW_HEIGHT,
            max(
                _wrapped_lines_for_width(p.name, _POL_COL_WIDTHS["C"]),
                _wrapped_lines_for_width(p.description, _POL_COL_WIDTHS["K"]),
            )
            * _LINE_HEIGHT_PT
            + 6,
        )

    ws.print_area = f"A1:{get_column_letter(_POL_OUT_COLS)}{row}"
    ws.print_title_rows = "1:5"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.print_options.horizontalCentered = True
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"


def _running_index(
    by_bucket: dict[str, list[Task]],
    bucket_names: list[str],
    current_bucket: str,
    within_idx: int,
) -> int:
    """Return the global 1-based item number across buckets in render order."""
    n = 0
    for name in bucket_names:
        if name == current_bucket:
            return n + within_idx
        n += len(by_bucket[name])
    return n + within_idx
