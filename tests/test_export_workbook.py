from __future__ import annotations

from datetime import date, timedelta

from openpyxl import load_workbook

from openitems.domain import checklists, engagements, policies, tasks
from openitems.domain.policies import PolicyInput
from openitems.domain.tasks import TaskInput
from openitems.export.theme import (
    CLR_RED,
)
from openitems.export.workbook import OUT_COLS, export_engagement


def _seed(session, today: date):
    e = engagements.create(session, "Acme")
    t1 = tasks.create(
        session,
        e,
        TaskInput(
            name="Migrate auth flow",
            description="Replace legacy session cookie with token-based auth.",
            priority="Urgent",
            status="In Progress",
            assigned_to="j.kowalski",
            start_date=today - timedelta(days=14),
            due_date=today - timedelta(days=3),  # overdue
            labels=["api", "sec"],
            bucket_name="Engineering",
        ),
    )
    checklists.add(session, t1, "spike token issuer", completed=True)
    checklists.add(session, t1, "draft RFC", completed=True)
    checklists.add(session, t1, "cutover comms", completed=False)

    tasks.create(
        session,
        e,
        TaskInput(
            name="Refactor exporter sheet",
            priority="Medium",
            status="In Progress",
            due_date=today + timedelta(days=10),
            labels=["api"],
            bucket_name="Engineering",
        ),
    )
    tasks.create(
        session,
        e,
        TaskInput(
            name="Doc: env var matrix",
            priority="Low",
            status="Intake",
            due_date=today + timedelta(days=30),
            labels=["docs"],
            bucket_name="Design",
        ),
    )
    # "Closed" is one of the seeded workflow buckets and is a done-state.
    tasks.create(
        session,
        e,
        TaskInput(
            name="Old completed thing",
            priority="Low",
            bucket_name="Closed",
        ),
    )
    # A Resolved task — also a done-state, also excluded from the export.
    tasks.create(
        session,
        e,
        TaskInput(
            name="Held in resolved",
            priority="Low",
            bucket_name="Resolved",
        ),
    )
    return e


def test_export_creates_workbook_with_expected_structure(session, tmp_path):
    today = date(2026, 5, 1)
    e = _seed(session, today)
    out = tmp_path / "open-items.xlsx"
    export_engagement(e, e.tasks, out, today=today)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active

    # Row 1 — client / engagement name
    assert ws["A1"].value == "  Acme"
    assert ws["A1"].font.bold is True

    # Row 2 — title
    assert ws["A2"].value == "  Open Items List"
    assert ws["A2"].font.bold is True

    # Row 3 — subtitle mentions counts (3 open, 2 buckets — Engineering + Design)
    assert "3 open items" in ws["A3"].value
    assert "2 buckets" in ws["A3"].value

    # Header row 5
    assert ws.cell(row=5, column=2).value == "#"
    assert ws.cell(row=5, column=3).value == "Task"
    assert ws.cell(row=5, column=9).value == "Description / Checklist"

    # First bucket header — workflow order: Engineering & Design are custom
    # (created by the seed via get_or_create after the seeded workflow stages),
    # so they come after the seeded stages but in creation order: Engineering
    # is mentioned first by the seed.
    bucket_headers = [
        ws.cell(row=r, column=1).value
        for r in range(6, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
        and ws.cell(row=r, column=1).value.strip().isupper()
        and ws.cell(row=r, column=2).value is None
    ]
    assert "  ENGINEERING" in bucket_headers
    assert "  DESIGN" in bucket_headers

    # Find the overdue task and confirm red+bold due-date
    overdue_row = None
    for row in ws.iter_rows(min_row=6, max_col=OUT_COLS):
        if row[2].value == "Migrate auth flow":
            overdue_row = row[0].row
            break
    assert overdue_row is not None
    due_cell = ws.cell(row=overdue_row, column=8)
    assert due_cell.font.color.rgb.endswith(CLR_RED)
    assert due_cell.font.bold is True

    # Priority cell for Urgent should be red+bold
    pri_cell = ws.cell(row=overdue_row, column=5)
    assert pri_cell.value == "Urgent"
    assert pri_cell.font.bold is True

    # Checklist rows present after overdue task
    label_cell = ws.cell(row=overdue_row + 1, column=3).value
    assert label_cell is not None and "Checklist:" in label_cell
    assert "[x]" in (ws.cell(row=overdue_row + 2, column=3).value or "")
    assert "[  ]" in (ws.cell(row=overdue_row + 4, column=3).value or "")

    # Print setup applied
    assert ws.print_area is not None
    assert ws.print_title_rows == "$1:$5"
    assert ws.freeze_panes == "A6"
    assert ws.sheet_view.showGridLines is False

    # Summary footer is somewhere later, mentions overdue count
    summary_text = None
    for row in ws.iter_rows(min_row=overdue_row + 6, max_col=1):
        v = row[0].value
        if v and "Summary:" in v:
            summary_text = v
            break
    assert summary_text is not None
    assert "1 overdue" in summary_text
    assert "1 high/urgent priority" in summary_text


def test_export_excludes_completed_and_deleted(session, tmp_path):
    today = date(2026, 5, 1)
    e = _seed(session, today)
    # Soft-delete one of the live tasks
    live = next(t for t in e.tasks if t.name == "Refactor exporter sheet")
    tasks.soft_delete(session, live)
    session.flush()

    out = tmp_path / "open-items.xlsx"
    export_engagement(e, e.tasks, out, today=today)
    wb = load_workbook(out)
    ws = wb.active

    names = {ws.cell(row=r, column=3).value for r in range(6, ws.max_row + 1)}
    assert "Old completed thing" not in names
    assert "Refactor exporter sheet" not in names
    assert "Migrate auth flow" in names


def test_export_omits_policies_tab_when_no_policies(session, tmp_path):
    today = date(2026, 5, 1)
    e = _seed(session, today)
    out = tmp_path / "open-items.xlsx"
    export_engagement(e, e.tasks, out, today=today)
    wb = load_workbook(out)
    assert len(wb.sheetnames) == 1


def test_export_includes_policies_tab_when_present(session, tmp_path):
    today = date(2026, 5, 1)
    e = _seed(session, today)

    # Three policies: lapsed, soon-renewing, distant.
    policies.create(
        session,
        e,
        PolicyInput(
            name="GL 2025",
            carrier="Travelers",
            coverage="GL",
            policy_number="GL-9001",
            effective_date=today - timedelta(days=400),
            expiration_date=today - timedelta(days=10),  # lapsed
            location="HQ",
            description="Primary GL — needs renewal",
        ),
    )
    policies.create(
        session,
        e,
        PolicyInput(
            name="Auto 2026",
            carrier="Liberty Mutual",
            coverage="Auto",
            policy_number="A-44",
            effective_date=today - timedelta(days=120),
            expiration_date=today + timedelta(days=20),  # soon
            location="Fleet",
        ),
    )
    policies.create(
        session,
        e,
        PolicyInput(
            name="Umbrella",
            carrier="Chubb",
            coverage="Umbrella",
            effective_date=today - timedelta(days=120),
            expiration_date=today + timedelta(days=300),  # far
        ),
    )
    all_policies = policies.list_for(session, e)

    out = tmp_path / "open-items.xlsx"
    export_engagement(e, e.tasks, out, today=today, policies=all_policies)
    wb = load_workbook(out)
    assert len(wb.sheetnames) == 2
    pol_sheet_name = next(n for n in wb.sheetnames if n.startswith("Policies"))
    ws = wb[pol_sheet_name]

    # Title block
    assert ws["A1"].value == "  Acme"
    assert ws["A2"].value == "  Policies"
    assert "3 policies" in ws["A3"].value
    assert "1 lapsed" in ws["A3"].value
    assert "1 renewing ≤30d" in ws["A3"].value

    # Column header row
    assert ws.cell(row=5, column=3).value == "Policy"
    assert ws.cell(row=5, column=9).value == "Expiration"
    assert ws.cell(row=5, column=10).value == "Days to Renewal"

    # Sort: lapsed → soon → far
    name_col = [
        ws.cell(row=r, column=3).value
        for r in range(6, 9)
    ]
    assert name_col == ["GL 2025", "Auto 2026", "Umbrella"]

    # Lapsed row: expiration cell is red+bold; days cell shows lapsed marker.
    lapsed_exp = ws.cell(row=6, column=9)
    assert lapsed_exp.font.color.rgb.endswith(CLR_RED)
    assert lapsed_exp.font.bold is True
    assert "lapsed" in (ws.cell(row=6, column=10).value or "")

    # Soon row: expiration cell is red but NOT bold (the quieter warning tier).
    soon_exp = ws.cell(row=7, column=9)
    assert soon_exp.font.color.rgb.endswith(CLR_RED)
    assert soon_exp.font.bold is False

    # Far row: expiration cell is plain.
    far_exp = ws.cell(row=8, column=9)
    assert far_exp.font.bold is False
    assert not far_exp.font.color.rgb.endswith(CLR_RED)

    # Sheet print setup
    assert ws.print_title_rows == "$1:$5"
    assert ws.freeze_panes == "A6"
    assert ws.sheet_view.showGridLines is False


def test_export_skips_deleted_policies(session, tmp_path):
    today = date(2026, 5, 1)
    e = _seed(session, today)
    p_keep = policies.create(
        session,
        e,
        PolicyInput(
            name="Keep",
            carrier="Travelers",
            coverage="GL",
            effective_date=today,
            expiration_date=today + timedelta(days=200),
        ),
    )
    p_drop = policies.create(
        session,
        e,
        PolicyInput(
            name="Drop",
            carrier="Travelers",
            coverage="GL",
            effective_date=today,
            expiration_date=today + timedelta(days=200),
        ),
    )
    policies.soft_delete(session, p_drop)
    session.flush()
    all_policies = policies.list_for(session, e, include_deleted=True)
    out = tmp_path / "open-items.xlsx"
    export_engagement(e, e.tasks, out, today=today, policies=all_policies)
    wb = load_workbook(out)
    pol_sheet_name = next(n for n in wb.sheetnames if n.startswith("Policies"))
    ws = wb[pol_sheet_name]
    names = {
        ws.cell(row=r, column=3).value
        for r in range(6, ws.max_row + 1)
        if ws.cell(row=r, column=3).value
    }
    assert names == {p_keep.name}
